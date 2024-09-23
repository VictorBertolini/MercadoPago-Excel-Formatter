from PyPDF2 import PdfReader
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from os import path, remove
from Nodes import *


# A function to get a pdf file and scrap all the text and put in a txt file 
def pdf_to_txt(pdf_file_name, txt_file_name = "Extrato.txt"):
    # Test if the pdf file exists
    if not path.isfile(pdf_file_name):
        print("[ Invalid pdf file ]")
        return -1

    # Open the pdf 
    with open(pdf_file_name, "rb") as pdf_file:
        # Open the txt file
        with open(txt_file_name, "w", encoding = "utf-8") as txt_file:

            # Read the pdf
            pdf_reader = PdfReader(pdf_file)
            
            # For each page, extract the text and write in txt file
            for i in range(len(pdf_reader.pages)):
                page = pdf_reader.pages[i]
                text = page.extract_text()
            
                txt_file.write(text + '\n')
    return 0

# Get all the text inside of a txt file and return as a list of lines
def getTxtLines(txt_file_name = "Extrato.txt"):
    # Test if the file exists
    if not path.isfile(txt_file_name):
        print("<< Coudn't reach to txt file >>")
        return None

    # Open the txt file and return the text 
    with open(txt_file_name, "r", encoding = "utf-8") as txt_file:
        return txt_file.readlines()

# Loop all the lines and find every unwanted sentences to remove from text
def remove_unwanted_sentences(unwanted_sentences, text):
    i = 0
    # See every line and remove each one has a unwanted sentence
    while i < len(text):
        for key_phrase in unwanted_sentences:
            if key_phrase in text[i]:
                text.pop(i)
                i -= 1
        i += 1

# Find all the commas in a string and return the positions
def find_commas(book):
    positions = []
    for i, char in enumerate(book):
        if char == ',':
            positions.append(i)

    return positions

# Transform all the lines of the list of lines (text)
# And make it a single string without '\n'
def text_to_book(text):
    sep = ""
    book = sep.join(text)
    book = book.replace('\n', ' ')
    return book

# Get all the lines and strip all and return the same list.strip()
def strip_lines(lines):
    new_lines = []
    for single_line in lines:
        new_lines.append(single_line.strip())
    
    return new_lines

# With a giant string, this function cut all lines looking in 
# 0000,00 the comma to add 3 positions and make a line 
# it observes 2 ',' and in the second one it jump 3 spaces and save as a line 
def construct_statement_lines(text):
    statement_lines = []
    book = text_to_book(text)
    comma_position = find_commas(book)
    init_position = 0

    # For each line there is 2 lines and i always want the second one
    for i in range(len(comma_position)):
        if i % 2 == 0:
            continue
        
        # Get the line final position (comma position + 3)
        line_final_pos = comma_position[i] + 3
        # Append a string with init and final position
        statement_lines.append(book[init_position:line_final_pos])
        # now the final position is the initial position
        init_position = line_final_pos
    
    # Strip all the lines and return the list of lines
    statement_lines = strip_lines(statement_lines)
    return statement_lines

# Save a list of lines in a txt file
def save_in_txt(statement_lines, txt_file_name = "Extrato.txt"):
    with open(txt_file_name, "w", encoding = "utf-8") as txt_file:
        for line in statement_lines:
            txt_file.write(line + "\n")

# Transform a list of lines in a node list
def text_to_node(text):
    node_list = []
    
    # Create a node and put inside it the information of the line of the text
    for line in text:
        node = Node()
        node.get_data(line)
        node_list.append(node)
    
    return node_list

# Function to cut the name of the people of cheap operations
def clean_text(list_key_operations, node_list, max_value_operation = 50):
    for node in node_list:
        for key_op in list_key_operations:
            if key_op in node.operation and node.value <= max_value_operation:
                node.cut_operation_code(key_op.strip())

# Save the node list in a xlsx file 
def save_xlsx(node_list, excel_file_name = "Planilha Extrato.xlsx", sheet_name = "Extrato", money_entry = "ENTRADA", money_exit = "SAÍDA"):

    wb = openpyxl.Workbook()
    wb.create_sheet(sheet_name, 0)
    
    page = wb[sheet_name]

    page.append(["Data", "Type", "Operation", "Value"])

    for node in node_list:
        option = money_entry
        if node.is_negative == True:
            option = money_exit
        
        node.replace_comma_dot()

        page.append([f"{node.data}", f"{option}", f"{node.operation.strip()}", f"{node.value}"])

    wb.save(excel_file_name)

# styling the xlsx file
def xlsx_template(excel_file_name = "Planilha Extrato.xlsx", sheet_name = "Extrato"):
    
    if not path.isfile(excel_file_name):
        print("[ Invalid pdf file ]")
        return -1
    

    # Open the excel file
    wb = load_workbook(excel_file_name)
    ws = wb.active

    # Define the column width
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 11
    ws.column_dimensions['C'].width = 70
    ws.column_dimensions['D'].width = 15

    # define the row height dimension
    ws.row_dimensions[1].height = 30

    # Define the colors of the parts of the excel file
    fill_header = PatternFill(start_color = "00143c", end_color = "00143c", fill_type = "solid")
    fill_left   = PatternFill(start_color = "00143c", end_color = "00143c", fill_type = "solid")
    fill_file   = PatternFill(start_color = "FFC000", end_color = "FFC000", fill_type = "solid")

    # Define the font and styles of the parts of the excel file
    font_header = Font(name = 'Verdana', size = 14, bold = True, color = 'ffffff')
    font_left = Font(name = 'Verdana', size = 11, color = 'ffffff')
    font_file = Font(name = 'Verdana', size = 11, color = '000000')

    # Define the border
    border = Border(
        left=Side(border_style="thin", color="000000"),  
        right=Side(border_style="thin", color="000000"), 
        top=Side(border_style="thin", color="000000"),  
        bottom=Side(border_style="thin", color="000000")  
    )   

    # Define the alignment
    alignment = Alignment(horizontal = 'center', vertical = 'center')


    # Loop in all the column giving the style and formating 
    # Formating the A column
    for row in ws.iter_rows(min_col=1, max_col=1, min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.fill = fill_left  
            cell.font = font_left 
            cell.alignment = alignment

    # Formating all the file 
    for row in ws.iter_rows(min_col=2, max_col=4, min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.fill = fill_file
            cell.font = font_file  
            cell.border = border
    
    # Formating the Header 
    header = ['A1', 'B1', 'C1', 'D1']

    for cell in header:
        ws[cell].fill = fill_header
        ws[cell].font = font_header
        ws[cell].alignment = alignment

    # Save the changes
    wb.save(excel_file_name)


# Delete the txt file when program finished
def delete_txt(txt_file_name = "Extrato.txt"):
    if path.isfile(txt_file_name):
        remove(txt_file_name)



# Attach all the information in the excel file 
def attach_operations(node_list, key_operations_list):
    
    new_node_list = []
    for key_op in key_operations_list:
        node = Node()
        node.operation = key_op
        new_node_list.append(node)


    # For each node in node_list
    for i in range(len(node_list)):
        found_key_operation_node = 0
        position = -1
        # For each key_operation_list test if is equal or not
        for j in range(len(key_operations_list)):
            if node_list[i].operation is key_operations_list[j]:
                new_node_list[j].value += node_list[i].value
                found_key_operation_node = 1
                if node_list[i].is_negative == True:
                    new_node_list[j].is_negative = True
                break

        if found_key_operation_node == 0:
            new_node_list.append(node_list[i])
    
    return new_node_list



def get_bank_statement_to_excel_file(pdf_file_name, remove_lines_list, key_operations_list, attach_operations_name = True, save_txt = False, excel_file_name = "Planilha Extrato.xlsx", txt_file_name = "Extrato.txt", max_value_operation = 50, sheet_name = "Extrato", money_entry = "ENTRADA", money_exit = "SAÍDA"):
    
    # Read the pdf and put the information in a txt file 
    pdf_to_txt(pdf_file_name, txt_file_name)

    # Create a list of the txt lines
    text = getTxtLines(txt_file_name)

    # Remove all the lines with sentences don't wanted of text 
    # (This function creates a single paragraph of all the lines)
    remove_unwanted_sentences(remove_lines_list, text)

    # From a single paragraph, separate in lines again (now just with the information of the operations)
    statement_lines = construct_statement_lines(text)

    # Save all the lines in txt file again
    save_in_txt(statement_lines, txt_file_name)

    # Get all the lines of the txt file to 'text'
    text = getTxtLines(txt_file_name)

    # Encapsulate all the lines of the 'text' in Nodes 
    node_list = text_to_node(text)

    # Cut the name of the person in the operation and remain just the operation
    clean_text(key_operations_list, node_list, max_value_operation)

    # Attach All the operations equal
    if attach_operations_name:
        node_list = attach_operations(node_list, key_operations_list)

    # Get the nodes and create a xlsx file with theese information
    save_xlsx(node_list, excel_file_name, sheet_name, money_entry, money_exit)

    # Give Style to a excel file 
    xlsx_template(excel_file_name, sheet_name)

    # Delete the txt file
    if save_txt == False:
        delete_txt()

    return